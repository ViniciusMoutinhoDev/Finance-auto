"""
exporter.py — Exportação para Excel (.xlsx)
Gera um arquivo Excel formatado com duas abas: transações e relatório por categoria.
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def export_to_excel(transactions: list[dict]) -> bytes:
    """
    Cria um .xlsx com:
      - Aba 1: Todas as transações com formatação condicional (débito = vermelho, crédito = verde)
      - Aba 2: Resumo por categoria
    """
    wb = Workbook()

    # --- Aba 1: Transações ---
    ws_tx = wb.active
    ws_tx.title = "Transações"

    headers = ["Data", "Descrição", "Categoria", "Valor (R$)", "Tipo"]
    _write_header_row(ws_tx, headers)

    for tx in sorted(transactions, key=lambda x: x["date"]):
        row = [
            tx["date"],
            tx["description"],
            tx.get("category", "Outros"),
            abs(tx["amount"]),
            "Crédito" if tx["amount"] > 0 else "Débito",
        ]
        ws_tx.append(row)

        # Coloriza a linha pelo tipo
        last_row = ws_tx.max_row
        fill_color = "C8E6C9" if tx["amount"] > 0 else "FFCDD2"  # verde claro / vermelho claro
        for col in range(1, 6):
            ws_tx.cell(row=last_row, column=col).fill = PatternFill(
                start_color=fill_color, end_color=fill_color, fill_type="solid"
            )

    # Ajusta largura das colunas
    _auto_fit_columns(ws_tx, [12, 45, 20, 15, 10])

    # --- Aba 2: Resumo por Categoria ---
    ws_cat = wb.create_sheet("Por Categoria")
    _write_header_row(ws_cat, ["Categoria", "Total Gasto (R$)", "Qtd. Transações"])

    # Agrega os dados
    from collections import defaultdict
    cat_totals: dict[str, dict] = defaultdict(lambda: {"total": 0.0, "count": 0})
    for tx in transactions:
        if tx["amount"] < 0:
            cat = tx.get("category", "Outros")
            cat_totals[cat]["total"] += abs(tx["amount"])
            cat_totals[cat]["count"] += 1

    for cat, data in sorted(cat_totals.items(), key=lambda x: x[1]["total"], reverse=True):
        ws_cat.append([cat, round(data["total"], 2), data["count"]])

    _auto_fit_columns(ws_cat, [25, 20, 20])

    # Serializa para bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _write_header_row(ws, headers: list[str]) -> None:
    """Escreve a linha de cabeçalho com estilo."""
    ws.append(headers)
    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")


def _auto_fit_columns(ws, widths: list[int]) -> None:
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
